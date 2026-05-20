"""합본 자동화 메인 스크립트.

Usage:
    python main.py --source <폴더> --output <파일.xlsx> --project <프로젝트명>
                   [--round <차수>] [--optical EN|CN|NONE] [--api-key <key>]
"""

import argparse
import os
import sys
import io
import re
from pathlib import Path

# Windows CP949 터미널에서 한/중 문자 출력 보장
if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
if sys.stderr.encoding and sys.stderr.encoding.lower() not in ("utf-8", "utf8"):
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

import openpyxl

from gemini_client import GeminiClient
from processor import process_sheet
from assembler import build_hapbon


def main():
    parser = argparse.ArgumentParser(description="합본 자동화 도구 (Gemini 2.5 Flash)")
    parser.add_argument("--api-key", default=os.environ.get("GEMINI_API_KEY"), help="Gemini API 키")
    parser.add_argument("--source", required=True, help="원본 xlsx 폴더 경로")
    parser.add_argument("--output", required=True, help="출력 합본 경로 (.xlsx)")
    parser.add_argument("--project", required=True, help="프로젝트명")
    parser.add_argument("--round", default="", help="차수 (예: 3.3차)")
    parser.add_argument("--optical", default="EN", choices=["EN", "CN", "NONE"], help="기준 옵티컬 언어")
    parser.add_argument("--record", default="KR", choices=["KR", "EN", "JP"], help="녹음 언어")
    args = parser.parse_args()

    if not args.api_key:
        print("ERROR: Gemini API 키 필요. --api-key 또는 환경변수 GEMINI_API_KEY 설정")
        sys.exit(1)

    _init_cache(args.output)
    _cache = _load_cache()
    client = GeminiClient(args.api_key)
    project_title = f"{args.project} {args.round}".strip()

    print(f"\n{'='*60}")
    print(f"  합본 자동화 시작")
    print(f"  프로젝트: {project_title}")
    print(f"  소스: {args.source}")
    print(f"  출력: {args.output}")
    print(f"  옵티컬: {args.optical}")
    print(f"{'='*60}\n")

    # 1. 파일 수집
    xlsx_files = _scan_xlsx(args.source)
    if not xlsx_files:
        print("ERROR: xlsx 파일 없음")
        sys.exit(1)
    print(f"[수집] {len(xlsx_files)}개 파일:")
    for f in xlsx_files:
        print(f"  · {f.name}")
    print()

    # 2. 분류
    classifications = []  # {file, fpath, sheet, cls}
    log_entries = []

    for fpath in xlsx_files:
        print(f"[분류] {fpath.name}")
        try:
            wb = openpyxl.load_workbook(str(fpath), data_only=True, read_only=True)
        except Exception as e:
            _log(log_entries, fpath.name, "-", f"파일 열기 실패: {e}")
            continue

        for sname in wb.sheetnames:
            # 숨김 시트 스킵
            try:
                if wb[sname].sheet_state in ("hidden", "veryHidden"):
                    print(f"  [{sname}] 숨김 → 스킵")
                    continue
            except Exception:
                pass

            preview = _load_preview(wb[sname], 15)
            # 녹음 언어 빠른 체크
            if not _has_record_lang(preview, args.record):
                print(f"  [{sname}] 녹음 언어({args.record}) 없음 → 스킵")
                _log(log_entries, fpath.name, sname, f"녹음 언어({args.record}) 대사 없음")
                continue

            ckey = _cache_key(fpath.name, sname)
            if ckey in _cache:
                cls = _cache[ckey]
                print(f"  [{sname}] (캐시) {cls['type']}")
                classifications.append({
                    "file": fpath.name,
                    "fpath": str(fpath),
                    "sheet": sname,
                    "cls": cls,
                })
                continue
            print(f"  [{sname}] 분류 중...", end=" ", flush=True)
            try:
                cls = client.classify_sheet(fpath.name, sname, preview, args.optical)
                if cls["type"] == "SKIP":
                    reason = cls.get("skip_reason") or "분류 불가"
                    print(f"SKIP ({reason})")
                    _log(log_entries, fpath.name, sname, reason)
                else:
                    print(cls["type"])
                    _cache[ckey] = cls
                    _save_cache(_cache)
                    classifications.append({
                        "file": fpath.name,
                        "fpath": str(fpath),
                        "sheet": sname,
                        "cls": cls,
                    })
            except Exception as e:
                print(f"ERROR: {e}")
                _log(log_entries, fpath.name, sname, f"분류 오류: {e}")

        wb.close()

    print(f"\n[분류 완료] {len(classifications)}개 시트 처리 예정\n")

    # 3. 처리
    processed = []
    battle_groups: dict[str, list] = {}  # fpath → [item, ...]

    for item in classifications:
        if item["cls"]["type"] == "Type_전투":
            battle_groups.setdefault(item["fpath"], []).append(item)

    done_battle_files: set[str] = set()

    for item in classifications:
        fpath = item["fpath"]
        sname = item["sheet"]
        cls = item["cls"]
        fname = item["file"]
        t = cls["type"]

        print(f"[처리] {fname} / [{sname}] ({t})", end=" ")

        try:
            if t == "Type_명방캐릭터":
                out_name = _make_sheet_name(fname, sname, cls)
                processed.append({
                    "sheet_name_out": out_name,
                    "type": t,
                    "rows": None,
                    "source_fpath": fpath,
                    "source_sheet": sname,
                    "cls": cls,
                })
                print("→ 원본 복사 예정")

            elif t == "Type_전투":
                if fpath in done_battle_files:
                    print("→ 이미 처리됨(합산)")
                    continue
                done_battle_files.add(fpath)

                battle_items = battle_groups[fpath]
                # boss를 앞으로 정렬
                battle_items.sort(key=lambda x: (0 if "boss" in x["sheet"].lower() else 1))

                wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
                all_rows = []
                for bi in battle_items:
                    ws = wb[bi["sheet"]]
                    rows = process_sheet(ws, bi["cls"])
                    all_rows.extend(rows)
                    all_rows.append(None)  # 구분선
                if all_rows and all_rows[-1] is None:
                    all_rows.pop()  # 마지막 구분선 제거
                wb.close()

                out_name = "전투"
                processed.append({
                    "sheet_name_out": out_name,
                    "type": t,
                    "rows": all_rows,
                    "source_file": fname,
                    "cls": cls,
                })
                print(f"→ {len([r for r in all_rows if r is not None])}행 (합산)")

            elif t == "Type_짧은음성":
                # 병합셀 처리: read_only=False 필요
                wb = openpyxl.load_workbook(fpath, data_only=True)
                ws = wb[sname]
                rows = process_sheet(ws, cls)
                wb.close()
                out_name = _make_sheet_name(fname, sname, cls)
                processed.append({
                    "sheet_name_out": out_name,
                    "type": t,
                    "rows": rows,
                    "source_file": fname,
                    "cls": cls,
                })
                print(f"→ {len(rows)}행")

            else:
                wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
                ws = wb[sname]
                rows = process_sheet(ws, cls)
                wb.close()
                # PV 시트: 처리된 행에서 대표 캐릭터명 추출 → 단일캐릭터 PV만 접두사 적용
                if t == "Type_PV" and not cls.get("char_name_kr"):
                    dominant = _dominant_char(rows)
                    ratio = _dominant_char_ratio(rows, dominant)
                    if dominant and ratio >= 0.3:  # 30% 이상이어야 단일캐릭터 PV
                        cls = dict(cls)
                        cls["char_name_kr"] = dominant
                out_name = _make_sheet_name(fname, sname, cls)
                processed.append({
                    "sheet_name_out": out_name,
                    "type": t,
                    "rows": rows,
                    "source_file": fname,
                    "cls": cls,
                })
                print(f"→ {len(rows)}행")

        except Exception as e:
            import traceback
            print(f"\n  ERROR: {e}")
            traceback.print_exc()
            _log(log_entries, fname, sname, f"처리 오류: {e}")

    # 4. 조립
    print(f"\n[조립] 합본 생성 중...")
    processed = _deduplicate_sheet_names(processed)
    # 출력 경로 보정: 폴더이거나 .xlsx 없으면 자동 수정
    output = args.output
    if Path(output).is_dir():
        stem = f"{project_title}_합본.xlsx".replace(" ", "_")
        output = str(Path(output) / stem)
    elif not output.lower().endswith(".xlsx"):
        output += ".xlsx"
    os.makedirs(Path(output).parent, exist_ok=True)
    build_hapbon(processed, log_entries, project_title, output, args.optical)
    print(f"[완료] {output}\n")
    print(f"[완료] {args.output}\n")


# ── 헬퍼 ─────────────────────────────────────────────────────────────────

def _dominant_char(rows: list[dict]) -> str | None:
    """행 목록에서 가장 많이 등장하는 캐릭터명 반환."""
    from collections import Counter
    counts = Counter(
        r.get("캐릭터명", "") for r in rows
        if r and r.get("캐릭터명") and r.get("캐릭터명") != "내레이션"
    )
    return counts.most_common(1)[0][0] if counts else None


def _dominant_char_ratio(rows: list[dict], char: str | None) -> float:
    """해당 캐릭터가 전체 행에서 차지하는 비율."""
    if not char or not rows:
        return 0.0
    total = sum(1 for r in rows if r and r.get("캐릭터명"))
    match = sum(1 for r in rows if r and r.get("캐릭터명") == char)
    return match / total if total > 0 else 0.0


def _deduplicate_sheet_names(processed: list[dict]) -> list[dict]:
    """출력 시트명 중복 시 _2, _3 접미사로 구분."""
    seen: dict[str, int] = {}
    for item in processed:
        name = item["sheet_name_out"]
        key = name.lower()
        if key in seen:
            seen[key] += 1
            item["sheet_name_out"] = f"{name}_{seen[key]}"[:31]
        else:
            seen[key] = 1
    return processed


def _scan_xlsx(folder: str) -> list[Path]:
    p = Path(folder)
    files = sorted(p.rglob("*.xlsx"))
    return [f for f in files if not f.name.startswith("~$")]


def _load_preview(ws, max_rows: int = 15) -> list[list]:
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        rows.append(list(row))
    return rows


def _has_record_lang(preview: list[list], record: str) -> bool:
    """녹음 언어에 해당하는 텍스트가 시트 미리보기에 존재하는지 확인."""
    for row in preview:
        for cell in row:
            if not cell:
                continue
            s = str(cell)
            if record == "KR" and any("가" <= c <= "힣" for c in s):
                return True
            elif record == "JP" and any("ぁ" <= c <= "ん" or "ァ" <= c <= "ン" for c in s):
                return True
            elif record == "EN" and re.search(r"[A-Za-z]{5,}", s):
                return True
    return False


def _log(log_entries: list, file: str, sheet: str, reason: str):
    log_entries.append({"file": file, "sheet": sheet, "reason": reason})


def _make_sheet_name(fname: str, sname: str, cls: dict) -> str:
    """출력 시트명 결정."""
    char_name = cls.get("char_name_kr") or ""
    sheet_type = cls.get("type", "")

    def strip_cn(s: str) -> str:
        s = re.sub(r"[一-鿿㐀-䶿豈-﫿぀-ヿ]+", "", s)
        s = re.sub(r"\b\d{6,}\b", "", s)
        s = re.sub(r"[_\-\s.]+", "_", s).strip("_").strip()
        return s

    type_labels = {
        "Type_메인": "메인", "Type_PV": "버전PV", "Type_전투": "전투",
        "Type_캐릭터음성": "음성", "Type_짧은음성": "짧은음성", "Type_무한대": "무한대",
    }

    if re.fullmatch(r"Sheet\d*", sname, re.IGNORECASE):
        stem = strip_cn(Path(fname).stem)
        kr_words = re.findall(r"[가-힣]{2,}", stem)
        en_parts = [p for p in re.findall(r"[A-Za-z]{3,}", stem) if p.lower() not in ("ver", "the", "for")]
        if kr_words:
            cleaned = "_".join(kr_words[:2])
        elif en_parts:
            cleaned = "_".join(en_parts[:2])
        else:
            cleaned = type_labels.get(sheet_type, "시트")
    else:
        cleaned = strip_cn(sname)
        if not cleaned:
            cleaned = type_labels.get(sheet_type, "시트")

    if char_name and char_name not in cleaned:
        cleaned = f"{char_name}_{cleaned}"

    if not re.search(r"[가-힣a-zA-Z]", cleaned):
        cleaned = type_labels.get(sheet_type, "기타")

    cleaned = re.sub(r"[:/\?*\[\]]", "", cleaned)
    return cleaned[:31]



import hashlib
import json as _json_mod

_CACHE_FILE = None

def _init_cache(output_path: str):
    global _CACHE_FILE
    from pathlib import Path
    _CACHE_FILE = Path(output_path).parent / ".hapbon_cls_cache.json"

def _cache_key(fname: str, sname: str) -> str:
    return hashlib.md5(f"{fname}||{sname}".encode()).hexdigest()

def _load_cache() -> dict:
    if _CACHE_FILE and _CACHE_FILE.exists():
        try:
            return _json_mod.loads(_CACHE_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}

def _save_cache(cache: dict):
    if _CACHE_FILE:
        _CACHE_FILE.write_text(_json_mod.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
