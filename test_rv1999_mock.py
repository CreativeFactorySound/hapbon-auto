"""리버스1999 모의(Mock) 분류 결과로 전체 파이프라인 테스트.

API 키 없이도 processor + assembler 전체를 검증합니다.
"""

import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

from pathlib import Path
import openpyxl

from processor import process_sheet, scan_gray_rows
from assembler import build_hapbon
from profile import load_profile, make_sheet_name_rv1999

SRC = Path(r"V:\DB_Work_202503\01_Game\HAOPLAY\20260520_리버스 1999_3.7차(24차)\자료\3.7 대본 20260521")
OUT = SRC / "리버스1999_3.7차_합본_모의테스트.xlsx"

PROFILE = load_profile("reverse1999")

# ── 수동 분류 결과 (Gemini 없이) ─────────────────────────────────────
MOCK_CLS = {
    # 캐릭터음성 파일들
    "SP 이글 음성 3.7SP小春雀儿角色语音_260427.xlsx": {
        "Sheet1": {
            "type": "Type_캐릭터음성",
            "header_row": 0,
            "col_filename": 10,
            "col_char_rec": -1, "col_char_cn": -1,
            "col_emotion_cn": 3, "col_emotion_rec": 4,
            "col_dialogue_cn": 6, "col_dialogue_rec": 8,
            "col_optical": -1, "optical_used": "EN",
            "col_alt": 9, "col_duration": -1, "col_step_type": -1,
            "col_functional": 1,
            "col_adr_wild": -1, "col_timecode": -1,
            "skip_step_values": [],
            "char_name_kr": "이글",
            "battle_role": "unknown", "notes": "", "skip_reason": None,
        }
    },
    "노티카 스킨 3.7诺谛卡皮肤语音_260506.xlsx": {
        "Sheet1": {
            "type": "Type_캐릭터음성",
            "header_row": 0,
            "col_filename": 8,
            "col_char_rec": -1, "col_char_cn": -1,
            "col_emotion_cn": 3, "col_emotion_rec": 4,
            "col_dialogue_cn": 6, "col_dialogue_rec": 7,
            "col_optical": -1, "optical_used": "EN",
            "col_alt": -1, "col_duration": -1, "col_step_type": -1,
            "col_functional": 1,
            "col_adr_wild": -1, "col_timecode": -1,
            "skip_step_values": [],
            "char_name_kr": "노티카",
            "battle_role": "unknown", "notes": "", "skip_reason": None,
        }
    },
    "리아논 음성 3.7小瑞安侬角色语音_260427.xlsx": {
        "Sheet1": {
            "type": "Type_캐릭터음성",
            "header_row": 0,
            "col_filename": 11,
            "col_char_rec": -1, "col_char_cn": -1,
            "col_emotion_cn": 3, "col_emotion_rec": 4,
            "col_dialogue_cn": 6, "col_dialogue_rec": 8,
            "col_optical": -1, "optical_used": "EN",
            "col_alt": 9, "col_duration": -1, "col_step_type": -1,
            "col_functional": 1,
            "col_adr_wild": -1, "col_timecode": -1,
            "skip_step_values": [],
            "char_name_kr": "리아논",
            "battle_role": "unknown", "notes": "", "skip_reason": None,
        }
    },
    "미스 스트레인저 음성 3.7无名者角色语音_260427.xlsx": {
        "Sheet1": {
            "type": "Type_캐릭터음성",
            "header_row": 0,
            "col_filename": 11,
            "col_char_rec": -1, "col_char_cn": -1,
            "col_emotion_cn": 3, "col_emotion_rec": 4,
            "col_dialogue_cn": 6, "col_dialogue_rec": 8,
            "col_optical": -1, "optical_used": "EN",
            "col_alt": -1, "col_duration": -1, "col_step_type": -1,
            "col_functional": 1,
            "col_adr_wild": -1, "col_timecode": -1,
            "skip_step_values": [],
            "char_name_kr": "미스 스트레인저",
            "battle_role": "unknown", "notes": "", "skip_reason": None,
        }
    },
    "코펠리아 음성 3.7珐琅眼角色语音_260427.xlsx": {
        "Sheet1": {
            "type": "Type_캐릭터음성",
            "header_row": 0,
            "col_filename": 11,
            "col_char_rec": -1, "col_char_cn": -1,
            "col_emotion_cn": 3, "col_emotion_rec": 4,
            "col_dialogue_cn": 6, "col_dialogue_rec": 8,
            "col_optical": -1, "optical_used": "EN",
            "col_alt": 9, "col_duration": -1, "col_step_type": -1,
            "col_functional": 1,
            "col_adr_wild": -1, "col_timecode": -1,
            "skip_step_values": [],
            "char_name_kr": "코펠리아",
            "battle_role": "unknown", "notes": "", "skip_reason": None,
        }
    },
    "코펠리아 스킨 3.7珐琅眼皮肤语音_260427.xlsx": {
        "Sheet1": {
            "type": "Type_캐릭터음성",
            "header_row": 0,
            "col_filename": 10,
            "col_char_rec": -1, "col_char_cn": -1,
            "col_emotion_cn": 3, "col_emotion_rec": 4,
            "col_dialogue_cn": 6, "col_dialogue_rec": 8,
            "col_optical": -1, "optical_used": "EN",
            "col_alt": -1, "col_duration": -1, "col_step_type": -1,
            "col_functional": 1,
            "col_adr_wild": -1, "col_timecode": -1,
            "skip_step_values": [],
            "char_name_kr": "코펠리아",
            "battle_role": "unknown", "notes": "", "skip_reason": None,
        }
    },
    "파투투 스킨 3.7图图石子皮肤语音_260427.xlsx": {
        "Sheet1": {
            "type": "Type_캐릭터음성",
            "header_row": 0,
            "col_filename": 9,
            "col_char_rec": -1, "col_char_cn": -1,
            "col_emotion_cn": 3, "col_emotion_rec": 4,
            "col_dialogue_cn": 6, "col_dialogue_rec": 7,
            "col_optical": -1, "optical_used": "EN",
            "col_alt": -1, "col_duration": -1, "col_step_type": -1,
            "col_functional": 1,
            "col_adr_wild": -1, "col_timecode": -1,
            "skip_step_values": [],
            "char_name_kr": "파투투",
            "battle_role": "unknown", "notes": "", "skip_reason": None,
        }
    },
    # PV
    "리아논 PV 3.7小瑞安侬角色PV_260210_260427.xlsx": {
        "Sheet1": {
            "type": "Type_PV",
            "header_row": 0,
            "col_filename": 1,
            "col_char_rec": 3, "col_char_cn": 2,
            "col_emotion_cn": 4, "col_emotion_rec": 5,
            "col_dialogue_cn": 7, "col_dialogue_rec": 8,
            "col_optical": 10, "optical_used": "EN",
            "col_alt": 9, "col_duration": 6, "col_step_type": -1,
            "col_functional": -1,
            "col_adr_wild": -1, "col_timecode": -1,
            "skip_step_values": [],
            "char_name_kr": "리아논",
            "battle_role": "unknown", "notes": "", "skip_reason": None,
        }
    },
    "버전 PV 3.7版本PV_260210_260427.xlsx": {
        "Sheet1": {
            "type": "Type_PV",
            "header_row": 0,
            "col_filename": 1,
            "col_char_rec": 3, "col_char_cn": 2,
            "col_emotion_cn": 4, "col_emotion_rec": 5,
            "col_dialogue_cn": 7, "col_dialogue_rec": 8,
            "col_optical": 10, "optical_used": "EN",
            "col_alt": 9, "col_duration": 6, "col_step_type": -1,
            "col_functional": -1,
            "col_adr_wild": -1, "col_timecode": -1,
            "skip_step_values": [],
            "char_name_kr": None,
            "battle_role": "unknown", "notes": "", "skip_reason": None,
        }
    },
    # 캐릭터 심층 분석 (Type_PV에 가까운 구조)
    "리아논 캐릭터 심층 분석 3.7小瑞安侬角色侧写_20260430.xlsx": {
        "Sheet1": {
            "type": "Type_PV",
            "header_row": 0,
            "col_filename": -1,
            "col_char_rec": 2, "col_char_cn": 1,
            "col_emotion_cn": 4, "col_emotion_rec": 5,
            "col_dialogue_cn": 3, "col_dialogue_rec": 7,
            "col_optical": 6, "optical_used": "EN",
            "col_alt": -1, "col_duration": -1, "col_step_type": -1,
            "col_functional": -1,
            "col_adr_wild": -1, "col_timecode": -1,
            "skip_step_values": [],
            "char_name_kr": "리아논",
            "battle_role": "unknown", "notes": "", "skip_reason": None,
        }
    },
    # 메인 대본
    "타인의 슬픔 3.7第十三章-他者的悲哀_20260521.xlsx": {
        "正文定稿": {
            "type": "Type_메인",
            "header_row": 0,
            "col_filename": 19,
            "col_char_rec": 6, "col_char_cn": 5,
            "col_emotion_cn": 16, "col_emotion_rec": 17,
            "col_dialogue_cn": 9, "col_dialogue_rec": 10,
            "col_optical": -1, "optical_used": "EN",
            "col_alt": -1, "col_duration": -1, "col_step_type": 4,
            "col_functional": -1,
            "col_adr_wild": -1, "col_timecode": -1,
            "skip_step_values": ["画面", "spine对话"],
            "char_name_kr": None,
            "battle_role": "unknown", "notes": "", "skip_reason": None,
        }
    },
    # 올로그
    "올로그 3.7永在律对话_20260521.xlsx": {
        "文本": {
            "type": "Type_메인",
            "header_row": 0,
            "col_filename": 9,
            "col_char_rec": 3, "col_char_cn": -1,
            "col_emotion_cn": 6, "col_emotion_rec": 7,
            "col_dialogue_cn": 4, "col_dialogue_rec": 5,
            "col_optical": -1, "optical_used": "EN",
            "col_alt": -1, "col_duration": -1, "col_step_type": -1,
            "col_functional": -1,
            "col_adr_wild": -1, "col_timecode": -1,
            "skip_step_values": [],
            "char_name_kr": None,
            "battle_role": "unknown", "notes": "", "skip_reason": None,
        }
    },
    # 전투
    "보스 및 전투 대화 3.7boss战斗对话&普通战中对话_260210_Latis_260427.xlsx": {
        "boss对话": {
            "type": "Type_전투",
            "header_row": 1,
            "col_filename": 14,
            "col_char_rec": 4, "col_char_cn": 3,
            "col_emotion_cn": 11, "col_emotion_rec": -1,
            "col_dialogue_cn": 7, "col_dialogue_rec": 8,  # KR 대사 없음 → -1
            "col_optical": -1, "optical_used": "EN",
            "col_alt": -1, "col_duration": -1, "col_step_type": 1,
            "col_functional": -1,
            "col_adr_wild": -1, "col_timecode": -1,
            "skip_step_values": [],
            "char_name_kr": None,
            "battle_role": "boss", "notes": "", "skip_reason": None,
        },
        "普通战斗对话": {
            "type": "Type_전투",
            "header_row": 2,
            "col_filename": 13,
            "col_char_rec": 4, "col_char_cn": 3,
            "col_emotion_cn": 11, "col_emotion_rec": -1,
            "col_dialogue_cn": 7, "col_dialogue_rec": -1,
            "col_optical": -1, "optical_used": "EN",
            "col_alt": -1, "col_duration": -1, "col_step_type": 1,
            "col_functional": -1,
            "col_adr_wild": -1, "col_timecode": -1,
            "skip_step_values": [],
            "char_name_kr": None,
            "battle_role": "normal", "notes": "", "skip_reason": None,
        },
    },
}


def _resolve_fpath(src: Path, fname: str) -> Path | None:
    """유니코드 정규화 차이 무시하고 실제 경로 반환."""
    import unicodedata
    fname_nfc = unicodedata.normalize("NFC", fname)
    fname_nfd = unicodedata.normalize("NFD", fname)
    for candidate in src.iterdir():
        cn = unicodedata.normalize("NFC", candidate.name)
        if cn == fname_nfc or cn == fname_nfd:
            return candidate
    return None


def run_mock_test():
    print("\n" + "="*55)
    print("  리버스1999 모의(Mock) 파이프라인 테스트")
    print("="*55 + "\n")

    classifications = []
    log_entries = []

    for fname, sheets in MOCK_CLS.items():
        fpath_resolved = _resolve_fpath(SRC, fname)
        if fpath_resolved is None:
            print(f"  [SKIP] 파일 없음: {fname}")
            continue
        fpath = fpath_resolved
        for sname, cls in sheets.items():
            cls = {
                **cls,
                "_record": "KR",
                "_functional_prefix": PROFILE.get("functional_emotion_prefix", True),
            }
            classifications.append({
                "file": fname,
                "fpath": str(fpath_resolved),
                "sheet": sname,
                "cls": cls,
            })
            print(f"  [분류] {fname[:40]} / {sname} → {cls['type']}")

    print(f"\n[처리 시작] {len(classifications)}개 시트\n")

    processed = []
    battle_groups: dict = {}
    for item in classifications:
        if item["cls"]["type"] == "Type_전투":
            battle_groups.setdefault(item["fpath"], []).append(item)
    done_battle: set = set()

    for item in classifications:
        fpath = item["fpath"]
        sname = item["sheet"]
        cls = item["cls"]
        fname = item["file"]
        t = cls["type"]
        print(f"  [처리] {fname[:35]} / {sname} ({t})", end=" ")

        try:
            if t == "Type_전투":
                if fpath in done_battle:
                    print("→ 합산됨")
                    continue
                done_battle.add(fpath)
                items = battle_groups[fpath]
                items.sort(key=lambda x: (0 if x["cls"].get("battle_role") == "boss" else 1))
                wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
                all_rows = []
                for bi in items:
                    ws = wb[bi["sheet"]]
                    rows = process_sheet(ws, bi["cls"])
                    all_rows.extend(rows)
                    all_rows.append(None)
                if all_rows and all_rows[-1] is None:
                    all_rows.pop()
                wb.close()
                processed.append({
                    "sheet_name_out": "전투",
                    "type": t,
                    "rows": all_rows,
                    "source_file": fname,
                    "cls": cls,
                })
                print(f"→ {len([r for r in all_rows if r])}행")

            elif t == "Type_짧은음성":
                gray = scan_gray_rows(fpath, sname, cls.get("col_dialogue_rec", -1), cls.get("header_row", 0))
                if gray:
                    cls = {**cls, "_gray_rows": gray}
                wb = openpyxl.load_workbook(fpath, data_only=True)
                ws = wb[sname]
                rows = process_sheet(ws, cls)
                wb.close()
                out_name = make_sheet_name_rv1999(fname, sname, cls)
                processed.append({
                    "sheet_name_out": out_name,
                    "type": t,
                    "rows": rows,
                    "source_file": fname,
                    "cls": cls,
                })
                print(f"→ {len(rows)}행")

            else:
                gray = scan_gray_rows(fpath, sname, cls.get("col_dialogue_rec", -1), cls.get("header_row", 0))
                if gray:
                    cls = {**cls, "_gray_rows": gray}
                wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
                ws = wb[sname]
                rows = process_sheet(ws, cls)
                wb.close()
                out_name = make_sheet_name_rv1999(fname, sname, cls)
                processed.append({
                    "sheet_name_out": out_name,
                    "type": t,
                    "rows": rows,
                    "source_file": fname,
                    "cls": cls,
                })
                print(f"→ {len(rows)}행")

        except Exception as e:
            import traceback
            print(f"\n  ERROR: {e}")
            traceback.print_exc()
            log_entries.append({"file": fname, "sheet": sname, "reason": str(e)})

    # 탭명 중복 제거
    seen: dict[str, int] = {}
    for item in processed:
        k = item["sheet_name_out"].lower()
        if k in seen:
            seen[k] += 1
            item["sheet_name_out"] = f"{item['sheet_name_out']}_{seen[k]}"[:31]
        else:
            seen[k] = 1

    print(f"\n[조립] 합본 생성 중...")
    build_hapbon(
        processed, log_entries,
        "리버스1999 3.7차",
        str(OUT), "EN",
        summary_mode="both",
        include_summary=True,
    )

    # 결과 검증
    wb_out = openpyxl.load_workbook(str(OUT), read_only=True, data_only=True)
    print(f"\n[결과 검증] 시트 목록:")
    for s in wb_out.sheetnames:
        ws = wb_out[s]
        nrows = sum(1 for _ in ws.iter_rows())
        print(f"  · {s}  ({nrows}행)")

    # 감정열 검증 (functional prefix 없어야 함)
    print("\n[감정 검증] 이글 (음성) 첫 5행:")
    if "이글 (음성)" in wb_out.sheetnames:
        ws = wb_out["이글 (음성)"]
        rows = list(ws.iter_rows(max_row=6, values_only=True))
        for i, row in enumerate(rows):
            emo = row[2] if len(row) > 2 else ""
            dial = str(row[4])[:30] if len(row) > 4 else ""
            print(f"  row{i}: 감정={str(emo)[:40]}  /  대사={dial}")
    else:
        print("  이글 (음성) 시트 없음")

    wb_out.close()
    print(f"\n[완료] → {OUT}\n")

    total_lines = sum(
        len([r for r in (item.get("rows") or []) if r is not None])
        for item in processed
    )
    print(f"총 {total_lines}라인\n")


if __name__ == "__main__":
    run_mock_test()
