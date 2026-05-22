"""합본 xlsx 조립 및 포맷 적용."""

import re
from pathlib import Path

# 지문·이펙트 제거 후 어절 수 계산 (개괄 단어 수 집계용)
_STAGE_DIR_RE = re.compile(r'\([^)]*\)|\[[^\]]*\]')

def _count_words(text: str) -> int:
    """괄호 안 지문을 제거한 뒤 공백 기준 어절 수 반환."""
    if not text:
        return 0
    cleaned = _STAGE_DIR_RE.sub('', str(text)).strip()
    return len(cleaned.split()) if cleaned else 0
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


# ── 색상 상수 ────────────────────────────────────────────────────────────────
HDR_BG = "1F3864"
HDR_FG = "FFFFFF"
DIAL_BG = "FFFDE7"
CHAR_BG = "92D050"
REC_BG  = "EA9999"
EMO_BG  = "FFE599"
ROW_ALT = "F5F5F5"
TIMING_BG = "FFE5CC"
PEAK_FG = "FF0000"
SUMMARY_BG_HDR = "1F3864"
SUMMARY_FG_HDR = "FFFFFF"
SUMMARY_TOTAL_BG = "D9E1F2"
SUMMARY_CAST_BG = "FFF2CC"

# 테두리
_THIN   = Side(style="thin",   color="BFBFBF")
_MEDIUM = Side(style="medium", color="999999")
_CELL_BORDER  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HDR_BORDER   = Border(left=_MEDIUM, right=_MEDIUM, top=_MEDIUM, bottom=_MEDIUM)

# 탭 그룹 순서 (template_spec.md 기준)
GROUP_ORDER = ["메인", "전투", "PV", "스킬", "캐릭터대본", "짧은음성", "음성", "스킨"]

TYPE_GROUP = {
    "Type_메인": "메인",
    "Type_전투": "전투",
    "Type_PV": "PV",
    "Type_캐릭터음성": "음성",
    "Type_짧은음성": "짧은음성",
    "Type_무한대": "메인",
    "Type_명방캐릭터": "캐릭터대본",
}


def build_hapbon(
    processed: list[dict],
    log_entries: list[dict],
    project_title: str,
    output_path: str,
    optical_lang: str,
):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    # 그룹별 정렬
    grouped = _group_sheets(processed)

    no_optical = (optical_lang == "NONE")

    # 데이터 시트 추가
    all_sheet_data = []  # [(sheet_name, rows)] — 개괄 집계용
    for group in GROUP_ORDER:
        for item in grouped.get(group, []):
            sname = item["sheet_name_out"][:31]
            if item["type"] == "Type_명방캐릭터":
                _add_arknights_sheet(wb, item)
            else:
                rows = item.get("rows") or []
                _add_data_sheet(wb, sname, rows, item["type"], images=item.get("images"), no_optical=no_optical)
                all_sheet_data.append((sname, rows, item["type"]))

    # 개괄 시트 (맨 앞)
    _add_summary_sheet(wb, all_sheet_data)

    # 로그 시트 (스킵된 항목 있을 때만)
    if log_entries:
        _add_log_sheet(wb, log_entries)

    # 시트 순서 재배치: 개괄 → 로그 → 데이터
    _reorder_sheets(wb)

    wb.save(output_path)


# ── 그룹핑 ─────────────────────────────────────────────────────────────────

def _group_sheets(processed: list[dict]) -> dict:
    groups: dict[str, list] = {g: [] for g in GROUP_ORDER}
    for item in processed:
        g = TYPE_GROUP.get(item["type"], "메인")
        groups[g].append(item)
    return groups


# ── 데이터 시트 ───────────────────────────────────────────────────────────

# ALT는 항상 포함 (숨김 처리 없음)
# 옵티컬(원문)·⏱ 검수는 optical_lang == "NONE"이면 제외
_BASE_COLS          = ["파일명", "캐릭터명", "감정", "REC", "대사", "옵티컬(원문)", "⏱ 검수", "ALT"]
_BASE_COLS_NO_OPT   = ["파일명", "캐릭터명", "감정", "REC", "대사", "ALT"]
_INFINITE_COLS      = ["파일명", "캐릭터명", "감정", "ADR/Wild", "타임코드", "REC", "대사", "옵티컬(원문)", "⏱ 검수", "ALT"]
_INFINITE_COLS_NO_OPT = ["파일명", "캐릭터명", "감정", "ADR/Wild", "타임코드", "REC", "대사", "ALT"]

# 중요 열 강조 폰트 크기
_FONT_LARGE = 14   # 캐릭터명, 대사, ALT
_FONT_SMALL = 9    # 나머지


def _get_col_list(sheet_type: str, no_optical: bool = False) -> list[str]:
    if sheet_type == "Type_무한대":
        return list(_INFINITE_COLS_NO_OPT if no_optical else _INFINITE_COLS)
    return list(_BASE_COLS_NO_OPT if no_optical else _BASE_COLS)


def _add_data_sheet(wb: openpyxl.Workbook, sheet_name: str, rows: list[dict], sheet_type: str,
                    images: list | None = None, no_optical: bool = False):
    ws = wb.create_sheet(title=sheet_name)
    if not rows:
        return

    cols = _get_col_list(sheet_type, no_optical=no_optical)

    # 원본 열 보존: 모든 행의 _extra_cols 키를 순서 유지하며 수집
    extra_col_names: list[str] = []
    seen_extra: set[str] = set()
    for row in rows:
        if row is None:
            continue
        for k in (row.get("_extra_cols") or {}):
            if k not in seen_extra:
                extra_col_names.append(k)
                seen_extra.add(k)

    all_cols = cols + extra_col_names

    # 헤더
    _write_header(ws, all_cols)

    # 데이터 행
    for r_idx, row in enumerate(rows):
        if row is None:  # 전투 합산 구분선
            ws.append([""] * len(all_cols))
            for c_idx in range(1, len(all_cols) + 1):
                ws.cell(row=r_idx + 2, column=c_idx).border = _CELL_BORDER
            continue
        excel_row = r_idx + 2
        is_even   = (r_idx % 2 == 0)
        is_peak   = row.get("_peak", False)
        is_timing = row.get("_timing_over", False)

        extra_dict = row.get("_extra_cols") or {}
        vals = [row.get(c, "") for c in cols] + [extra_dict.get(c, "") for c in extra_col_names]
        ws.append(vals)

        _format_data_row(ws, excel_row, all_cols, is_even, is_peak, is_timing)

    _set_col_widths(ws, all_cols)
    ws.freeze_panes = "A2"

    # 원본 열(extra) 숨김 처리
    for i in range(len(cols) + 1, len(all_cols) + 1):
        ws.column_dimensions[get_column_letter(i)].hidden = True

    # 이미지 삽입 (Type_짧은음성 등)
    if images:
        _insert_images(ws, rows, images, cols)


def _write_header(ws, cols: list[str]):
    ws.append(cols)
    hdr_fill  = PatternFill("solid", fgColor=HDR_BG)
    hdr_font  = Font(name="맑은 고딕", size=10, bold=True, color=HDR_FG)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = hdr_align
        cell.border    = _HDR_BORDER
    ws.row_dimensions[1].height = 25


def _format_data_row(ws, row_idx: int, cols: list[str], is_even: bool, is_peak: bool, is_timing: bool):
    base_bg = ROW_ALT if is_even else "FFFFFF"
    if is_timing:
        base_bg = TIMING_BG

    col_bg = {
        "파일명":    base_bg,
        "캐릭터명":  CHAR_BG,
        "감정":      EMO_BG,
        "REC":       REC_BG,
        "대사":      DIAL_BG,
        "ALT":       DIAL_BG,
        "옵티컬(원문)": base_bg,
        "⏱ 검수":   base_bg,
        "ADR/Wild":  EMO_BG,
        "타임코드":  base_bg,
    }

    for c_idx, col_name in enumerate(cols, 1):
        cell = ws.cell(row=row_idx, column=c_idx)
        bg   = col_bg.get(col_name, base_bg)

        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border    = _CELL_BORDER

        # 폰트 — 캐릭터명·대사·ALT는 14pt Bold
        is_char  = col_name == "캐릭터명"
        is_dial  = col_name in ("대사", "ALT")
        is_large = is_char or is_dial
        font_color = PEAK_FG if (is_peak and is_dial) else "000000"

        cell.font = Font(
            name="맑은 고딕",
            size=_FONT_LARGE if is_large else _FONT_SMALL,
            bold=is_large,
            color=font_color,
            italic=(col_name == "캐릭터명" and str(cell.value) == "내레이션"),
        )


def _set_col_widths(ws, cols: list[str]):
    widths = {
        "파일명":       40,
        "캐릭터명":     16,
        "감정":         22,
        "REC":          5,
        "대사":         55,
        "ALT":          45,
        "옵티컬(원문)": 45,
        "⏱ 검수":      18,
        "ADR/Wild":     10,
        "타임코드":     26,
    }
    for i, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 15)


# ── 이미지 삽입 (Type_짧은음성 등) ──────────────────────────────────────────

def _insert_images(ws, rows: list[dict], images: list[dict], cols: list[str]):
    """출력 시트의 데이터 열 오른쪽에 이미지를 캐릭터 섹션별로 배치.

    Pillow 미설치 시 gracefully skip.
    """
    import io as _io
    try:
        from openpyxl.drawing.image import Image as _XLImage
    except ImportError:
        return

    if not images:
        return

    # 캐릭터별 출력 행 번호 목록 (1-based, 헤더=1 → 데이터 시작=2)
    char_rows: dict[str, list[int]] = {}
    for r_idx, row in enumerate(rows):
        if row is None:
            continue
        char = (row.get("캐릭터명") or "").strip()
        if char:
            char_rows.setdefault(char, []).append(r_idx + 2)

    # 이미지 열: 마지막 데이터 열 오른쪽
    img_col_idx    = len(cols) + 1
    img_col_letter = get_column_letter(img_col_idx)
    ws.column_dimensions[img_col_letter].width = 36   # ≈ 270 px

    for img_info in images:
        raw       = img_info.get("bytes", b"")
        char_name = img_info.get("char_name", "")
        cx_emu    = img_info.get("cx_emu", 0)
        cy_emu    = img_info.get("cy_emu", 0)

        if not raw:
            continue

        # 캐릭터 섹션 행 찾기 (완전 매칭 → 부분 매칭 fallback)
        row_list = char_rows.get(char_name, [])
        if not row_list:
            for c, rl in char_rows.items():
                if char_name and (char_name in c or c in char_name):
                    row_list = rl
                    break
        if not row_list:
            row_list = [2]

        start_row = row_list[0]

        try:
            img_obj = _XLImage(_io.BytesIO(raw))

            # 픽셀 크기 결정 (96 DPI: 1 px = 9525 EMU)
            if cx_emu > 0 and cy_emu > 0:
                w_px = round(cx_emu / 9525)
                h_px = round(cy_emu / 9525)
            else:
                w_px = getattr(img_obj, 'width',  240) or 240
                h_px = getattr(img_obj, 'height', 320) or 320

            # 열 폭(270 px)에 맞게 비율 유지하며 축소, 세로 최대 450 px
            max_w, max_h = 270, 450
            scale = min(max_w / max(w_px, 1), max_h / max(h_px, 1), 1.0)
            w_px  = round(w_px * scale)
            h_px  = round(h_px * scale)

            img_obj.width  = max(w_px, 60)
            img_obj.height = max(h_px, 60)

            ws.add_image(img_obj, f"{img_col_letter}{start_row}")

        except Exception:
            pass


# ── 명방캐릭터 시트 (원본 그대로) ────────────────────────────────────────

def _add_arknights_sheet(wb: openpyxl.Workbook, item: dict):
    sname = item["sheet_name_out"][:31]
    ws = wb.create_sheet(title=sname)
    fpath = item.get("source_fpath")
    src_sheet = item.get("source_sheet")
    if not fpath or not src_sheet:
        return
    try:
        src_wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
        src_ws = src_wb[src_sheet]
        for row in src_ws.iter_rows(values_only=True):
            ws.append([v for v in row])
        src_wb.close()
    except Exception as e:
        ws.append([f"시트 복사 오류: {e}"])


# ── 개괄 시트 ──────────────────────────────────────────────────────────────

def _add_summary_sheet(wb: openpyxl.Workbook, all_sheet_data: list[tuple]):
    ws = wb.create_sheet(title="개괄")

    # 캐릭터별 탭별 라인 수 + 단어 수 집계
    char_tab_lines: dict[str, dict[str, int]] = {}
    char_tab_words: dict[str, dict[str, int]] = {}
    tab_names: list[str] = []

    _HEADER_CHAR_SKIP = {"캐릭터명", "캐릭터", "配音对象", "角色",
                         "キャラ", "キャラクター", "Character", "CHARACTER"}

    for sname, rows, stype in all_sheet_data:
        if stype == "Type_명방캐릭터":
            continue
        tab_names.append(sname)
        for row in (rows or []):
            if row is None:
                continue
            char = row.get("캐릭터명", "").strip()
            if not char or char in _HEADER_CHAR_SKIP:
                continue
            dial = str(row.get("대사") or "").strip()
            words = _count_words(dial)

            if char not in char_tab_lines:
                char_tab_lines[char] = {}
                char_tab_words[char] = {}
            char_tab_lines[char][sname] = char_tab_lines[char].get(sname, 0) + 1
            char_tab_words[char][sname] = char_tab_words[char].get(sname, 0) + words

    # 헤더: 탭 열에 "(라인/단어)" 부제목
    header = ["캐릭터명", "성우명"] + [f"{t}\n(라인/단어)" for t in tab_names] + ["합계\n(라인/단어)"]
    ws.append(header)

    # 헤더 스타일
    hdr_fill = PatternFill("solid", fgColor=SUMMARY_BG_HDR)
    hdr_font = Font(name="맑은 고딕", size=10, bold=True, color=SUMMARY_FG_HDR)
    for cell in ws[1]:
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _HDR_BORDER
    ws.row_dimensions[1].height = 36  # 2줄 헤더 높이

    # 캐릭터별 행 (총 라인 수 내림차순)
    sorted_chars = sorted(
        char_tab_lines.items(),
        key=lambda x: sum(x[1].values()),
        reverse=True,
    )
    for r_idx, (char, line_map) in enumerate(sorted_chars, 2):
        total_lines = sum(line_map.values())
        total_words = sum(char_tab_words[char].values())

        tab_cells = []
        for t in tab_names:
            ln = line_map.get(t, 0)
            wd = char_tab_words[char].get(t, 0)
            tab_cells.append(f"{ln} / {wd}" if ln else "")

        row_vals = [char, ""] + tab_cells + [f"{total_lines} / {total_words}"]
        ws.append(row_vals)

        for c_idx, _ in enumerate(row_vals, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            col_name = header[c_idx - 1]
            cell.border = _CELL_BORDER
            if col_name == "캐릭터명":
                cell.font = Font(name="맑은 고딕", size=9, bold=True)
            elif col_name == "성우명":
                cell.fill = PatternFill("solid", fgColor=SUMMARY_CAST_BG)
                cell.font = Font(name="맑은 고딕", size=9)
            elif col_name.startswith("합계"):
                cell.fill = PatternFill("solid", fgColor=SUMMARY_TOTAL_BG)
                cell.font = Font(name="맑은 고딕", size=9, bold=True)
            else:
                cell.font = Font(name="맑은 고딕", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 열 폭
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 14
    for i in range(3, len(header) + 1):
        # 탭 이름 길이 기준 (헤더에 "\n(라인/단어)" 붙어있으므로 원본 탭명 기준)
        tab_name = tab_names[i - 3] if i - 3 < len(tab_names) else "합계"
        ws.column_dimensions[get_column_letter(i)].width = max(16, len(tab_name) * 1.5)

    ws.freeze_panes = "A2"


# ── 로그 시트 ──────────────────────────────────────────────────────────────

def _add_log_sheet(wb: openpyxl.Workbook, log_entries: list[dict]):
    ws = wb.create_sheet(title="로그")
    ws.append(["파일명", "시트명", "사유"])
    hdr_fill = PatternFill("solid", fgColor=HDR_BG)
    hdr_font = Font(name="맑은 고딕", size=10, bold=True, color=HDR_FG)
    for cell in ws[1]:
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal="center")
        cell.border    = _HDR_BORDER
    for entry in log_entries:
        ws.append([entry.get("file", ""), entry.get("sheet", ""), entry.get("reason", "")])
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 50


# ── 시트 순서 재배치 ─────────────────────────────────────────────────────

def _reorder_sheets(wb: openpyxl.Workbook):
    names = wb.sheetnames
    front = []
    back  = []
    for n in names:
        if n in ("개괄", "로그"):
            front.append(n)
        else:
            back.append(n)
    ordered = ["개괄"] + ([n for n in front if n != "개괄"]) + back
    # 로그는 개괄 바로 다음
    if "로그" in ordered:
        ordered.remove("로그")
        ordered.insert(1, "로그")
    for i, name in enumerate(ordered):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=wb.sheetnames.index(name) * -1 + i)
