"""타입별 시트 데이터 추출."""

import io as _io
import re
import zipfile as _zipfile
from xml.etree import ElementTree as _ET
import openpyxl


def process_sheet(ws, cls: dict) -> list[dict]:
    """분류 결과(cls)에 따라 시트에서 행 데이터를 추출해 반환."""
    t = cls["type"]
    if t == "Type_메인":
        return _extract_main(ws, cls)
    elif t == "Type_PV":
        return _extract_pv(ws, cls)
    elif t == "Type_전투":
        return _extract_battle(ws, cls)
    elif t == "Type_캐릭터음성":
        return _extract_chara(ws, cls)
    elif t == "Type_짧은음성":
        return _extract_short(ws, cls)
    elif t == "Type_무한대":
        return _extract_infinite(ws, cls)
    else:
        return []


# ── 공통 헬퍼 ────────────────────────────────────────────────────────────────

def _cell(row_vals: list, idx: int):
    """안전하게 컬럼 값 반환. idx=-1이면 None."""
    if idx < 0 or idx >= len(row_vals):
        return None
    v = row_vals[idx]
    return str(v).strip() if v is not None else None


# 유효하지 않은 캐릭터명 — 녹음용으로 의미 없는 placeholder 값
_INVALID_CHAR_NAMES = {"???", "？？？"}


def _valid_char(name: str | None) -> str | None:
    """캐릭터명이 유효한 경우에만 반환. ???·？？？ 등 플레이스홀더면 None."""
    if not name:
        return None
    if name.strip() in _INVALID_CHAR_NAMES:
        return None
    return name


def _est_sec_kr(text: str) -> float:
    """한국어 글자 수로 예상 녹음 시간(초) 추정."""
    if not text:
        return 0.0
    return sum(1 for c in text if "가" <= c <= "힣") / 6.5


def _est_sec_en(text: str) -> float:
    if not text:
        return 0.0
    return len(str(text).split()) / 2.5


def _est_sec_jp(text: str) -> float:
    """일본어 글자 수로 예상 녹음 시간(초) 추정 (히라가나/가타카나 1모라, 한자 2모라 기준 5.5모라/초)."""
    if not text:
        return 0.0
    morae = 0
    for c in text:
        if "ぁ" <= c <= "ん" or "ァ" <= c <= "ン":   # 히라가나/가타카나
            morae += 1
        elif "一" <= c <= "鿿" or "㐀" <= c <= "䶿":  # 한자
            morae += 2
    return morae / 5.5 if morae else len(text.split()) / 2.5


def _parse_duration(s) -> float | None:
    """'≤3s', '2.5초', '3-4s' 등에서 숫자(초) 파싱."""
    if s is None:
        return None
    s = str(s).replace("≤", "").replace("≦", "").replace("约", "").strip()
    # 범위 표기 (3-4s) → 최댓값 사용
    m = re.search(r"([\d.]+)\s*[-~]\s*([\d.]+)", s)
    if m:
        return float(m.group(2))
    m = re.search(r"([\d.]+)\s*[sS초]", s)
    if m:
        return float(m.group(1))
    return None


_PEAK_WORDS: dict[str, list[str]] = {
    "KR": ["비명", "기합", "으아", "절규", "소리치", "폭발", "외치", "포효",
           "분노", "열받", "살려", "싫어", "제발", "도망", "그만", "아악", "으윽"],
    "JP": ["叫び", "絶叫", "悲鳴", "叫ぶ", "吼える", "激怒", "助けて", "嫌だ",
           "うわ", "きゃあ", "ぎゃあ", "くそ", "やめて", "逃げろ"],
    "EN": ["scream", "shriek", "roar", "howl", "help me", "stop", "no no"],
}


def _is_peak(text: str, record: str = "KR") -> bool:
    if not text:
        return False
    words = _PEAK_WORDS.get(record, _PEAK_WORDS["KR"])
    return "!!" in text or any(w in text for w in words)


def _timing_flag(dialogue: str, duration_raw, record: str = "KR") -> str:
    dur = _parse_duration(duration_raw)
    if dur is None:
        return ""
    if record == "JP":
        est = _est_sec_jp(dialogue or "")
    elif record == "EN":
        est = _est_sec_en(dialogue or "")
    else:
        est = _est_sec_kr(dialogue or "")
    if est > dur * 1.1:
        return f"⚠ 예상{est:.1f}s > 허용{dur}s"
    return ""


def _ws_rows(ws, header_row: int) -> list[list]:
    """워크시트에서 header_row 이후 데이터 행들을 반환."""
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > header_row:
            rows.append(list(row))
    return rows


def _get_header(ws, header_row: int) -> list[str]:
    """헤더 행 값을 문자열 리스트로 반환."""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == header_row:
            return [str(v).strip() if v is not None else "" for v in row]
    return []


def _collect_extra(row_vals: list, used: set, header: list) -> dict:
    """표준 매핑에 사용되지 않은 열의 값을 {헤더명: 값} 딕셔너리로 반환.
    값이 있는 열만 포함하며, 동일 헤더명 중복 시 _{col번호} 접미사로 구분."""
    extra: dict[str, str] = {}
    seen_headers: dict[str, int] = {}
    for i, v in enumerate(row_vals):
        if i in used:
            continue
        val_str = str(v).strip() if v is not None else ""
        if not val_str:
            continue
        h = header[i] if i < len(header) and header[i] else f"열{i + 1}"
        if h in seen_headers:
            h = f"{h}_{i + 1}"
        seen_headers[h] = i
        extra[h] = val_str
    return extra


# ── 회색 셀 감지 ─────────────────────────────────────────────────────────────

def _is_gray_cell(cell) -> bool:
    """셀의 폰트 색상 또는 배경 색상이 회색 계열인지 확인."""
    def _parse_rgb(rgb_str: str):
        s = rgb_str.upper().strip()
        if len(s) == 8:
            s = s[2:]   # alpha 접두사(FF) 제거
        if len(s) != 6:
            return None
        if s in ("000000", "FFFFFF", "00000000", "FFFFFFFF"):
            return None
        try:
            return int(s[0:2], 16), int(s[2:4], 16), int(s[4:6], 16)
        except Exception:
            return None

    def _is_gray_rgb(rgb_str: str, lo: int, hi: int) -> bool:
        rgb = _parse_rgb(rgb_str)
        if rgb is None:
            return False
        r, g, b = rgb
        # R ≈ G ≈ B (±20) 이고 lo ≤ R ≤ hi 이면 회색
        return lo <= r <= hi and abs(r - g) <= 20 and abs(g - b) <= 20

    try:
        fc = cell.font.color if cell.font else None
        if fc and fc.type == "rgb" and fc.rgb:
            if _is_gray_rgb(fc.rgb, 60, 200):   # 회색 폰트
                return True
    except Exception:
        pass

    try:
        fill = cell.fill
        if fill and fill.patternType not in (None, "none"):
            fg = fill.fgColor
            if fg and fg.type == "rgb" and fg.rgb:
                if _is_gray_rgb(fg.rgb, 140, 230):  # 밝은 회색 배경
                    return True
    except Exception:
        pass

    return False


def scan_gray_rows(fpath: str, sname: str, dialogue_col: int, header_row: int) -> set:
    """대사 열이 회색 처리된 행의 절대 인덱스(0-based) 집합 반환.
    read_only=False로 열어야 스타일 정보 접근 가능.
    오류 발생 시 빈 set 반환 (graceful fallback).
    """
    if dialogue_col < 0:
        return set()
    gray: set[int] = set()
    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
        ws = wb[sname]
        for r_idx, row in enumerate(ws.iter_rows()):
            if r_idx <= header_row:
                continue
            if dialogue_col < len(row) and _is_gray_cell(row[dialogue_col]):
                gray.add(r_idx)
        wb.close()
    except Exception:
        pass
    return gray


# 내레이션 스텝값 — 여기 포함된 값은 모두 "내레이션"으로 정규화
_NARRATION_STEPS = {
    "旁白",          # CN
    "내레이션",      # KR
    "나레이션",      # KR 오타 변형
    "ナレーション",  # JP
    "ナレ",          # JP 약칭
    "Narration",     # EN
    "NARRATION",
    "Narrator",
    "NARRATOR",
}

# 헤더 반복 행 감지용 — 녹음 언어별 헤더 문자열 포함
_HEADER_CHAR_VALUES  = {"캐릭터", "配音对象", "角色", "キャラ", "キャラクター", "Character", "CHARACTER", "Char"}
_HEADER_DIAL_VALUES  = {"라인", "对话文本", "终选语音", "台詞", "テキスト", "Line", "LINE", "Dialogue", "DIALOGUE", "Text"}
_HEADER_FILE_VALUES  = {"语音命名", "台詞番号", "ボイス名", "Filename", "File Name", "FILENAME"}

# 녹음 불필요 행 감지용 — 이 값이 포함된 행은 스킵
_SKIP_ROW_KEYWORDS = {
    # 중국어
    "无需配音", "怪物音效", "不需配音", "无配音", "无需录音", "不需要配音",
    # 한국어
    "더빙 불필요", "녹음 불필요", "번역 불필요", "더빙불필요", "녹음불필요",
    "번역불필요", "녹음없음", "더빙없음", "녹음 없음", "더빙 없음",
    # 일본어
    "収録不要", "録音不要",
}


# ── Type_메인 ──────────────────────────────────────────────────────────────

def _extract_main(ws, cls: dict) -> list[dict]:
    header_row = cls.get("header_row", 0)
    record     = cls.get("_record", "KR")
    cf = cls.get("col_filename", -1)
    ckr = cls.get("col_char_rec", -1)
    ckn = cls.get("col_char_cn", -1)
    ekr = cls.get("col_emotion_rec", -1)
    ecn = cls.get("col_emotion_cn", -1)
    dkr = cls.get("col_dialogue_rec", -1)
    dcn = cls.get("col_dialogue_cn", -1)
    opt = cls.get("col_optical", -1)
    alt = cls.get("col_alt", -1)
    step_col = cls.get("col_step_type", -1)
    skip_steps = set(cls.get("skip_step_values", []))

    used    = {c for c in [cf, ckr, ckn, ekr, ecn, dkr, dcn, opt, alt, step_col] if c >= 0}
    header  = _get_header(ws, header_row)
    gray_rows = cls.get("_gray_rows", set())

    result = []
    for abs_idx, row_vals in enumerate(_ws_rows(ws, header_row), start=header_row + 1):
        filename = _cell(row_vals, cf)
        char_rec = _cell(row_vals, ckr)
        char_cn = _cell(row_vals, ckn)
        dialogue_rec = _cell(row_vals, dkr)
        dialogue_cn = _cell(row_vals, dcn)
        emotion_rec = _cell(row_vals, ekr)
        emotion_cn = _cell(row_vals, ecn)
        optical = _cell(row_vals, opt)
        alt_val = _cell(row_vals, alt)
        step = _cell(row_vals, step_col)

        # 스킵: 씬 타입이 skip_steps에 해당
        if step and step in skip_steps:
            continue

        # 스킵: 헤더 반복 행
        if char_rec in _HEADER_CHAR_VALUES or char_cn in _HEADER_CHAR_VALUES:
            continue
        if dialogue_rec in _HEADER_DIAL_VALUES or dialogue_cn in _HEADER_DIAL_VALUES:
            continue

        # 스킵: 녹음 불필요 표기 행
        row_text = " ".join(str(v) for v in row_vals if v is not None)
        if any(kw in row_text for kw in _SKIP_ROW_KEYWORDS):
            continue

        # 스킵: 파일명도 없고 대사도 없고 캐릭터도 없는 완전 빈 행
        if not filename and not dialogue_rec and not dialogue_cn and not char_rec and not char_cn:
            continue

        # 스킵: 회색 처리된 행 (번역 없음 + 회색 = 완전 제외)
        if abs_idx in gray_rows and not dialogue_rec:
            continue

        # 내레이션 스텝 → "내레이션" 정규화
        if step and step in _NARRATION_STEPS:
            char_rec = "내레이션"

        # 녹음 대상 대사 없고 CN만 있는 경우
        if not dialogue_rec and dialogue_cn:
            emotion_rec = (emotion_rec or "") + "[번역 없음]"

        # 감정: 녹음언어 우선, 없으면 CN
        emotion = emotion_rec or emotion_cn or ""

        # ???는 녹음 의미 없는 표시명 — 실제 캐릭터명 열 우선, 둘 다 ???면 그대로 사용
        char = _valid_char(char_rec) or _valid_char(char_cn) or char_rec or char_cn or ""
        result.append({
            "파일명": filename or "",
            "캐릭터명": char,
            "감정": emotion,
            "REC": "",
            "대사": dialogue_rec or "",
            "ALT": alt_val or "",
            "옵티컬(원문)": optical or dialogue_cn or "",
            "⏱ 검수": "",
            "_peak": _is_peak(dialogue_rec or "", record),
            "_extra_cols": _collect_extra(row_vals, used, header),
        })
    return result


# ── Type_PV ────────────────────────────────────────────────────────────────

def _extract_pv(ws, cls: dict) -> list[dict]:
    header_row = cls.get("header_row", 0)
    record     = cls.get("_record", "KR")
    ckr = cls.get("col_char_rec", -1)
    ckn = cls.get("col_char_cn", -1)
    ekr = cls.get("col_emotion_rec", -1)
    ecn = cls.get("col_emotion_cn", -1)
    dkr = cls.get("col_dialogue_rec", -1)
    dcn = cls.get("col_dialogue_cn", -1)
    opt = cls.get("col_optical", -1)
    alt = cls.get("col_alt", -1)
    dur = cls.get("col_duration", -1)

    used   = {c for c in [ckr, ckn, ekr, ecn, dkr, dcn, opt, alt, dur] if c >= 0}
    header = _get_header(ws, header_row)
    gray_rows = cls.get("_gray_rows", set())

    result = []
    for abs_idx, row_vals in enumerate(_ws_rows(ws, header_row), start=header_row + 1):
        char_rec = _cell(row_vals, ckr)
        char_cn = _cell(row_vals, ckn)
        dialogue_rec = _cell(row_vals, dkr)
        dialogue_cn = _cell(row_vals, dcn)
        emotion_rec = _cell(row_vals, ekr)
        emotion_cn = _cell(row_vals, ecn)
        optical = _cell(row_vals, opt)
        alt_val = _cell(row_vals, alt)
        duration_raw = _cell(row_vals, dur)

        # 빈 행 스킵
        if not char_rec and not char_cn and not dialogue_rec and not dialogue_cn:
            continue

        # 스킵: 녹음 불필요 표기 행
        row_text = " ".join(str(v) for v in row_vals if v is not None)
        if any(kw in row_text for kw in _SKIP_ROW_KEYWORDS):
            continue

        # 스킵: 회색 처리된 행
        if abs_idx in gray_rows and not dialogue_rec:
            continue

        # 내레이션 정규화 (캐릭터 없고 대사만 있는 행)
        if not char_rec and not char_cn and (dialogue_rec or dialogue_cn):
            char_rec = "내레이션"

        emotion = emotion_rec or emotion_cn or ""
        timing = _timing_flag(dialogue_rec or "", duration_raw, record) if dur >= 0 else ""
        char = _valid_char(char_rec) or _valid_char(char_cn) or char_rec or char_cn or ""

        result.append({
            "파일명": "",
            "캐릭터명": char,
            "감정": emotion,
            "REC": "",
            "대사": dialogue_rec or "",
            "ALT": alt_val or "",
            "옵티컬(원문)": optical or dialogue_cn or "",
            "⏱ 검수": timing,
            "_peak": _is_peak(dialogue_rec or "", record),
            "_timing_over": bool(timing),
            "_extra_cols": _collect_extra(row_vals, used, header),
        })
    return result


# ── Type_전투 ──────────────────────────────────────────────────────────────

def _extract_battle(ws, cls: dict) -> list[dict]:
    header_row = cls.get("header_row", 0)
    record     = cls.get("_record", "KR")
    cf = cls.get("col_filename", -1)
    ckr = cls.get("col_char_rec", -1)
    ckn = cls.get("col_char_cn", -1)
    ekr = cls.get("col_emotion_rec", -1)
    ecn = cls.get("col_emotion_cn", -1)
    dkr = cls.get("col_dialogue_rec", -1)
    dcn = cls.get("col_dialogue_cn", -1)
    opt = cls.get("col_optical", -1)
    alt = cls.get("col_alt", -1)
    process_col = cls.get("col_step_type", -1)  # 流程节点

    used   = {c for c in [cf, ckr, ckn, ekr, ecn, dkr, dcn, opt, alt, process_col] if c >= 0}
    header = _get_header(ws, header_row)
    gray_rows = cls.get("_gray_rows", set())

    result = []
    for abs_idx, row_vals in enumerate(_ws_rows(ws, header_row), start=header_row + 1):
        filename = _cell(row_vals, cf)
        char_rec = _cell(row_vals, ckr)
        char_cn = _cell(row_vals, ckn)
        dialogue_rec = _cell(row_vals, dkr)
        dialogue_cn = _cell(row_vals, dcn)
        emotion_rec = _cell(row_vals, ekr)
        emotion_cn = _cell(row_vals, ecn)
        optical = _cell(row_vals, opt)
        alt_val = _cell(row_vals, alt)
        process_val = _cell(row_vals, process_col)

        # 파트 제목 행 스킵
        if process_val and not char_rec and not char_cn and not dialogue_rec and not dialogue_cn:
            continue

        # 헤더 반복 행 스킵
        if char_rec in _HEADER_CHAR_VALUES or char_cn in _HEADER_CHAR_VALUES:
            continue
        if dialogue_rec in _HEADER_DIAL_VALUES or dialogue_cn in _HEADER_DIAL_VALUES:
            continue

        # 스킵: 녹음 불필요 표기 행
        row_text = " ".join(str(v) for v in row_vals if v is not None)
        if any(kw in row_text for kw in _SKIP_ROW_KEYWORDS):
            continue

        # 완전 빈 행 스킵
        if not filename and not char_rec and not char_cn and not dialogue_rec and not dialogue_cn:
            continue

        # 스킵: 회색 처리된 행
        if abs_idx in gray_rows and not dialogue_rec:
            continue

        emotion = emotion_rec or emotion_cn or ""
        char = _valid_char(char_rec) or _valid_char(char_cn) or char_rec or char_cn or ""
        result.append({
            "파일명": filename or "",
            "캐릭터명": char,
            "감정": emotion,
            "REC": "",
            "대사": dialogue_rec or "",
            "ALT": alt_val or "",
            "옵티컬(원문)": optical or dialogue_cn or "",
            "⏱ 검수": "",
            "_peak": _is_peak(dialogue_rec or "", record),
            "_extra_cols": _collect_extra(row_vals, used, header),
        })
    return result


# ── Type_캐릭터음성 ────────────────────────────────────────────────────────

def _extract_chara(ws, cls: dict) -> list[dict]:
    header_row = cls.get("header_row", 0)
    record     = cls.get("_record", "KR")
    cf = cls.get("col_filename", -1)
    char_name = cls.get("char_name_kr") or ""
    ekr = cls.get("col_emotion_rec", -1)
    ecn = cls.get("col_emotion_cn", -1)
    dkr = cls.get("col_dialogue_rec", -1)
    dcn = cls.get("col_dialogue_cn", -1)
    opt = cls.get("col_optical", -1)
    alt = cls.get("col_alt", -1)
    func_col = cls.get("col_functional", -1)

    used   = {c for c in [cf, ekr, ecn, dkr, dcn, opt, alt, func_col] if c >= 0}
    header = _get_header(ws, header_row)
    gray_rows = cls.get("_gray_rows", set())

    result = []
    for abs_idx, row_vals in enumerate(_ws_rows(ws, header_row), start=header_row + 1):
        filename = _cell(row_vals, cf)
        dialogue_rec = _cell(row_vals, dkr)
        dialogue_cn = _cell(row_vals, dcn)
        emotion_rec = _cell(row_vals, ekr)
        emotion_cn = _cell(row_vals, ecn)
        optical = _cell(row_vals, opt)
        alt_val = _cell(row_vals, alt)
        functional = _cell(row_vals, func_col)

        # 빈 행 스킵
        if not filename and not dialogue_rec and not dialogue_cn:
            continue
        # 헤더 반복 행 스킵
        if dialogue_rec in _HEADER_DIAL_VALUES or filename in _HEADER_FILE_VALUES:
            continue

        # 스킵: 녹음 불필요 표기 행
        row_text = " ".join(str(v) for v in row_vals if v is not None)
        if any(kw in row_text for kw in _SKIP_ROW_KEYWORDS):
            continue

        # 스킵: 회색 처리된 행
        if abs_idx in gray_rows and not dialogue_rec:
            continue

        # 녹음 대상 대사 없고 원문만 있는 경우
        if not dialogue_rec and dialogue_cn:
            emotion_rec = (emotion_rec or "") + "[번역 없음]"

        emotion = emotion_rec or emotion_cn or ""
        # 功能을 감정 앞에 참고 정보로 붙임 (프로파일 설정에 따라)
        if functional and cls.get("_functional_prefix", True):
            emotion = f"[{functional}] {emotion}".strip()

        result.append({
            "파일명": filename or "",
            "캐릭터명": char_name,
            "감정": emotion,
            "REC": "",
            "대사": dialogue_rec or "",
            "ALT": alt_val or "",
            "옵티컬(원문)": optical or dialogue_cn or "",
            "⏱ 검수": "",
            "_peak": _is_peak(dialogue_rec or "", record),
            "_extra_cols": _collect_extra(row_vals, used, header),
        })
    return result


# ── Type_짧은음성 ──────────────────────────────────────────────────────────

def _extract_short(ws, cls: dict) -> list[dict]:
    """병합셀 처리가 필요하므로 ws는 read_only=False로 열어야 함.

    캐릭터명뿐 아니라 감정(詳細 등) 열도 병합셀이 있을 수 있어서
    모든 열에 대해 병합셀 값을 통합 관리한다.
    """
    header_row = cls.get("header_row", 0)
    record     = cls.get("_record", "KR")
    cf  = cls.get("col_filename",    -1)
    ckr = cls.get("col_char_rec",     -1)
    ckn = cls.get("col_char_cn",     -1)
    ekr = cls.get("col_emotion_rec",  -1)
    ecn = cls.get("col_emotion_cn",  -1)
    dkr = cls.get("col_dialogue_rec", -1)
    dcn = cls.get("col_dialogue_cn", -1)
    opt = cls.get("col_optical",     -1)
    alt = cls.get("col_alt",         -1)

    # 모든 열의 병합셀 값 맵: (row_0based, col_0based) → 값
    merged_map: dict[tuple[int, int], str] = {}
    try:
        for mr in ws.merged_cells.ranges:
            val = ws.cell(mr.min_row, mr.min_col).value
            s   = str(val).strip() if val is not None else ""
            c0  = mr.min_col - 1   # 0-based
            for r1 in range(mr.min_row, mr.max_row + 1):
                merged_map[(r1 - 1, c0)] = s  # 0-based row key
    except Exception:
        pass  # read_only 모드이면 merged_cells 접근 불가 → 그냥 진행

    def _mv(row_vals: list, row_0: int, col_0: int) -> str | None:
        """병합셀 우선으로 값 반환. col_0=-1이면 None."""
        if col_0 < 0:
            return None
        mv = merged_map.get((row_0, col_0))
        if mv is not None:
            return mv or None   # 빈 문자열은 None으로
        return _cell(row_vals, col_0)

    used   = {c for c in [cf, ckr, ckn, ekr, ecn, dkr, dcn, opt, alt] if c >= 0}
    header = _get_header(ws, header_row)
    gray_rows = cls.get("_gray_rows", set())

    result = []
    all_rows = list(ws.iter_rows(values_only=True))
    for i, row_vals in enumerate(all_rows):
        if i <= header_row:
            continue
        row_vals = list(row_vals)

        filename    = _mv(row_vals, i, cf)
        char_rec     = _mv(row_vals, i, ckr)
        char_cn     = _mv(row_vals, i, ckn)
        emotion_rec  = _mv(row_vals, i, ekr)
        emotion_cn  = _mv(row_vals, i, ecn)
        dialogue_rec = _mv(row_vals, i, dkr)
        dialogue_cn = _mv(row_vals, i, dcn)
        optical     = _mv(row_vals, i, opt)
        alt_val     = _mv(row_vals, i, alt)

        # 대사 없는 행 스킵
        if not dialogue_rec and not dialogue_cn:
            continue
        # 파일명·대사 둘 다 없는 행도 스킵
        if not filename and not dialogue_rec:
            continue

        # 스킵: 녹음 불필요 표기 행
        row_text = " ".join(str(v) for v in row_vals if v is not None)
        if any(kw in row_text for kw in _SKIP_ROW_KEYWORDS):
            continue

        # 스킵: 회색 처리된 행 (i = 절대 row 인덱스, 녹음 언어 대사 없으면 제외)
        if i in gray_rows and not dialogue_rec:
            continue

        if not dialogue_rec and dialogue_cn:
            emotion_rec = (emotion_rec or "") + "[번역 없음]"

        emotion = emotion_rec or emotion_cn or ""
        char = _valid_char(char_rec) or _valid_char(char_cn) or char_rec or char_cn or ""
        result.append({
            "파일명":       filename or "",
            "캐릭터명":     char,
            "감정":         emotion,
            "REC":          "",
            "대사":         dialogue_rec or "",
            "ALT":          alt_val or "",
            "옵티컬(원문)": optical or dialogue_cn or "",
            "⏱ 검수":      "",
            "_peak":        _is_peak(dialogue_rec or "", record),
            "_extra_cols":  _collect_extra(row_vals, used, header),
        })
    return result


# ── Type_무한대 ────────────────────────────────────────────────────────────

def _extract_infinite(ws, cls: dict) -> list[dict]:
    header_row = cls.get("header_row", 0)
    record     = cls.get("_record", "KR")
    cf = cls.get("col_filename", -1)
    ckr = cls.get("col_char_rec", -1)
    ckn = cls.get("col_char_cn", -1)
    dkr = cls.get("col_dialogue_rec", -1)
    dcn = cls.get("col_dialogue_cn", -1)
    opt = cls.get("col_optical", -1)
    adr_col = cls.get("col_adr_wild", -1)
    tc_col = cls.get("col_timecode", -1)
    rec_col = -1  # 녹음여부 컬럼은 헤더에서 탐색

    # 녹음여부 컬럼 찾기 (KR/CN/JP/EN 헤더 키워드 모두 처리)
    _REC_KEYWORDS = {"收录", "Record", "수록", "录音", "録音有無", "Include", "Rec"}
    if header_row >= 0:
        header_vals = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == header_row:
                header_vals = list(row)
                break
        for j, v in enumerate(header_vals):
            if v and any(kw in str(v) for kw in _REC_KEYWORDS):
                rec_col = j
                break

    used   = {c for c in [cf, ckr, ckn, dkr, dcn, opt, adr_col, tc_col, rec_col] if c >= 0}
    header = _get_header(ws, header_row)
    gray_rows = cls.get("_gray_rows", set())

    result = []
    for abs_idx, row_vals in enumerate(_ws_rows(ws, header_row), start=header_row + 1):
        # 녹음여부 필터
        if rec_col >= 0:
            rec_val = _cell(row_vals, rec_col)
            if rec_val and rec_val.lower() in ("no", "否", "n"):
                continue

        filename = _cell(row_vals, cf)
        char_rec = _cell(row_vals, ckr)
        char_cn = _cell(row_vals, ckn)
        dialogue_rec = _cell(row_vals, dkr)
        dialogue_cn = _cell(row_vals, dcn)
        optical = _cell(row_vals, opt)
        adr_wild = _cell(row_vals, adr_col)
        timecode = _cell(row_vals, tc_col)

        if not dialogue_rec and not dialogue_cn:
            continue

        # 스킵: 녹음 불필요 표기 행
        row_text = " ".join(str(v) for v in row_vals if v is not None)
        if any(kw in row_text for kw in _SKIP_ROW_KEYWORDS):
            continue

        # 스킵: 회색 처리된 행
        if abs_idx in gray_rows and not dialogue_rec:
            continue

        # 타임코드에서 duration 계산
        timing = ""
        if timecode and "-->" in str(timecode):
            dur = _parse_timecode_duration(str(timecode))
            if dur:
                timing = _timing_flag(dialogue_rec or "", f"{dur}s", record)

        char = _valid_char(char_rec) or _valid_char(char_cn) or char_rec or char_cn or ""
        result.append({
            "파일명": filename or "",
            "캐릭터명": char,
            "감정": "",
            "ADR/Wild": adr_wild or "",
            "타임코드": timecode or "",
            "REC": "",
            "대사": dialogue_rec or "",
            "ALT": "",
            "옵티컬(원문)": optical or dialogue_cn or "",
            "⏱ 검수": timing,
            "_peak": _is_peak(dialogue_rec or ""),
            "_timing_over": bool(timing),
            "_extra_cols": _collect_extra(row_vals, used, header),
        })
    return result


def _parse_timecode_duration(tc: str) -> float | None:
    """'00:00:05.700 --> 00:00:08.140' 형식에서 duration(초) 계산."""
    parts = tc.split("-->")
    if len(parts) != 2:
        return None
    try:
        def to_sec(s):
            s = s.strip().replace(",", ".")
            h, m, rest = s.split(":")
            return int(h) * 3600 + int(m) * 60 + float(rest)
        return to_sec(parts[1]) - to_sec(parts[0])
    except Exception:
        return None


# ── 이미지 추출 (zipfile 직접 파싱, twoCellAnchor 포함) ──────────────────────

def extract_sheet_images(fpath: str, sheet_name: str) -> list[dict]:
    """xlsx에서 지정 시트의 이미지를 추출한다 (openpyxl이 못 읽는 twoCellAnchor 포함).

    반환값: [{"bytes": bytes, "ext": str, "from_row": int, "from_col": int,
              "to_row": int, "to_col": int, "cx_emu": int, "cy_emu": int}, ...]
    row/col은 0-based.
    """
    NS_MAIN  = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    NS_PKG_R = 'http://schemas.openxmlformats.org/package/2006/relationships'
    NS_OFF_R = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
    NS_XDR   = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
    NS_A     = 'http://schemas.openxmlformats.org/drawingml/2006/main'

    result: list[dict] = []
    try:
        with _zipfile.ZipFile(fpath, 'r') as zf:
            names = set(zf.namelist())

            # 1. 시트 r:id 조회
            wb_root = _ET.fromstring(zf.read('xl/workbook.xml'))
            rid = None
            for sh in wb_root.iter(f'{{{NS_MAIN}}}sheet'):
                if sh.get('name') == sheet_name:
                    rid = sh.get(f'{{{NS_OFF_R}}}id')
                    break
            if not rid:
                return result

            # 2. 시트 파일 경로
            wb_rels = _ET.fromstring(zf.read('xl/_rels/workbook.xml.rels'))
            sheet_path = None
            for rel in wb_rels.iter(f'{{{NS_PKG_R}}}Relationship'):
                if rel.get('Id') == rid:
                    t = rel.get('Target', '')
                    sheet_path = ('xl/' + t) if not t.startswith('xl/') else t
                    break
            if not sheet_path or sheet_path not in names:
                return result

            # 3. sheet.xml.rels → drawing 경로
            sp = sheet_path.rsplit('/', 1)
            sheet_rels_path = f"{sp[0]}/_rels/{sp[1]}.rels"
            if sheet_rels_path not in names:
                return result

            sh_rels = _ET.fromstring(zf.read(sheet_rels_path))
            drawing_path = None
            for rel in sh_rels.iter(f'{{{NS_PKG_R}}}Relationship'):
                if 'drawing' in rel.get('Type', '').lower():
                    t = rel.get('Target', '')
                    if t.startswith('../'):
                        drawing_path = 'xl/' + t[3:]
                    elif t.startswith('/'):
                        drawing_path = t[1:]
                    else:
                        drawing_path = sp[0] + '/' + t
                    break
            if not drawing_path or drawing_path not in names:
                return result

            # 4. drawing.xml.rels → 이미지 경로 맵
            dp = drawing_path.rsplit('/', 1)
            dr_rels_path = f"{dp[0]}/_rels/{dp[1]}.rels"
            rid_to_media: dict[str, str] = {}
            if dr_rels_path in names:
                dr_rels = _ET.fromstring(zf.read(dr_rels_path))
                for rel in dr_rels.iter(f'{{{NS_PKG_R}}}Relationship'):
                    t = rel.get('Target', '')
                    if t.startswith('../'):
                        media_path = 'xl/' + t[3:]
                    elif t.startswith('/'):
                        media_path = t[1:]
                    else:
                        media_path = dp[0] + '/' + t
                    rid_to_media[rel.get('Id', '')] = media_path

            # 5. drawing XML → anchor 파싱
            dr_root = _ET.fromstring(zf.read(drawing_path))
            for tag in ('twoCellAnchor', 'oneCellAnchor'):
                for anchor in dr_root.findall(f'{{{NS_XDR}}}{tag}'):
                    from_el = anchor.find(f'{{{NS_XDR}}}from')
                    to_el   = anchor.find(f'{{{NS_XDR}}}to')
                    fr_row = int(from_el.find(f'{{{NS_XDR}}}row').text) if from_el is not None else 0
                    fr_col = int(from_el.find(f'{{{NS_XDR}}}col').text) if from_el is not None else 0
                    to_row = int(to_el.find(f'{{{NS_XDR}}}row').text)   if to_el   is not None else fr_row + 10
                    to_col = int(to_el.find(f'{{{NS_XDR}}}col').text)   if to_el   is not None else fr_col + 3

                    # 표시 크기 (EMU) — xdr:ext 또는 a:ext
                    ext_el = anchor.find(f'.//{{{NS_XDR}}}ext')
                    if ext_el is None:
                        ext_el = anchor.find(f'.//{{{NS_A}}}ext')
                    cx = int(ext_el.get('cx', 0)) if ext_el is not None else 0
                    cy = int(ext_el.get('cy', 0)) if ext_el is not None else 0

                    # blip r:embed
                    blip = anchor.find(f'.//{{{NS_A}}}blip')
                    if blip is None:
                        continue
                    r_embed = blip.get(f'{{{NS_OFF_R}}}embed', '')
                    if not r_embed or r_embed not in rid_to_media:
                        continue

                    media_path = rid_to_media[r_embed]
                    if media_path not in names:
                        continue

                    result.append({
                        "bytes":    zf.read(media_path),
                        "ext":      media_path.rsplit('.', 1)[-1].lower(),
                        "from_row": fr_row,   # 0-based
                        "from_col": fr_col,
                        "to_row":   to_row,
                        "to_col":   to_col,
                        "cx_emu":   cx,
                        "cy_emu":   cy,
                    })
    except Exception:
        pass

    return result


def extract_images_for_sheet(fpath: str, sname: str, cls: dict) -> list[dict]:
    """Type_짧은음성 시트 이미지를 추출하고 각 이미지에 캐릭터명(char_name)을 매핑한다."""
    images = extract_sheet_images(fpath, sname)
    if not images:
        return []

    ckr = cls.get("col_char_rec", -1)
    ckn = cls.get("col_char_cn", -1)

    # 병합셀 → row(0-based) → char 이름
    row_to_char: dict[int, str] = {}
    try:
        wb_tmp = openpyxl.load_workbook(fpath, data_only=True)
        ws_tmp = wb_tmp[sname]
        for mr in ws_tmp.merged_cells.ranges:
            c0 = mr.min_col - 1  # 0-based
            if c0 not in (ckr, ckn):
                continue
            val = ws_tmp.cell(mr.min_row, mr.min_col).value
            char = str(val).strip() if val else ""
            for r in range(mr.min_row - 1, mr.max_row):   # 0-based
                row_to_char[r] = char
        wb_tmp.close()
    except Exception:
        pass

    for img in images:
        char = ""
        for r in range(img["from_row"], img["to_row"] + 1):
            if r in row_to_char:
                char = row_to_char[r]
                break
        img["char_name"] = char

    return images
